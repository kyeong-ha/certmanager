# backend/api/cert/management/commands/import_certificates.py

import uuid
from openpyxl import load_workbook
from django.core.management.base import BaseCommand

# 관련 모델 import
from api.user.models.User import User
from api.cert.models.Certificate import Certificate
from api.center.models.EducationCenter import EducationCenter
from api.center.models.EducationCenterSession import EducationCenterSession

class Command(BaseCommand):
    help = 'Import certificate data from Excel file'
    file_path = './data/test_certificates.xlsx'
    
    # # 명령어 인자 정의: python manage.py import_certificates <excel_file>
    # def add_arguments(self, parser):
    #     parser.add_argument('excel_file', type=str)

    # 실제 실행 로직
    def handle(self, *args, **kwargs):
        # file_path = kwargs['excel_file']
        wb = load_workbook(self.file_path)
        ws = wb.active

        # 엑셀의 두 번째 행부터 데이터 반복 처리
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            (
                issue_number,         # 발급번호
                issue_date,           # 발급일자
                user_name,            # 사용자 이름
                birth_date,           # 생년월일
                course_name,          # 자격과정
                phone_number,         # 휴대폰 번호
                center_name,          # 교육원_기수 형태 문자열 (예: '교육원명_1기')
                user_id,              # 회원 ID
                address,              # 회원 주소
            ) = row[:9]

            # 1. 교육원명 / 기수 분리
            try:
                center_name, session_label = center_name.split("_")
                center_session = int(session_label.replace("기", ""))
            except Exception:
                self.stderr.write(f"[오류] {i}행 파싱 실패: {center_name}")
                continue

            # 2. 교육기관 조회 또는 생성
            center, _ = EducationCenter.objects.get_or_create(center_name=center_name)

            # 3. 이전 session에서 배송주소 상속
            delivery_address = ''
            if center_session > 1:
                prev_session = EducationCenterSession.objects.filter(
                    education_center=center,
                    center_session=center_session - 1
                ).first()
                if prev_session:
                    delivery_address = prev_session.delivery_address

            # 4. 세션 생성 또는 조회
            session, _ = EducationCenterSession.objects.get_or_create(
                education_center=center,
                center_session=center_session,
                defaults={'delivery_address': delivery_address}
            )

            # 5. 사용자 생성
            try:
                user = User.objects.get(phone_number=phone_number)
                updated = False
                if user.user_name != user_name:
                    user.user_name = user_name
                    updated = True
                if user.birth_date != birth_date:
                    user.birth_date = birth_date
                    updated = True
                if user.user_id != user_id:
                    user.user_id = user_id
                    updated = True
                if user.address != address:
                    user.address = address
                    updated = True
                if updated:
                    user.save()
                created = False
            except User.DoesNotExist:
                user = User.objects.create(
                    user_name=user_name,
                    birth_date=birth_date,
                    phone_number=phone_number,
                    user_id=user_id,
                    address=address,
                )
                created = True
                    
            # 6. 자격증 생성
            certificate = Certificate.objects.create(
                user=user,
                course_name=course_name,
                issue_number=issue_number,
                issue_date=issue_date,
                education_center=center,
                education_session=session,
                delivery_address=address,
                tracking_number=None,
            )

            # 7. 운송장 리스트에 등록
            if certificate.tracking_number:
                if certificate.tracking_number not in session.tracking_numbers:
                    session.tracking_numbers.append(certificate.tracking_number)
                    session.save()

            # 8. 로그 출력
            self.stdout.write(f"[{i}] {user_name} - {issue_number} 등록 완료")